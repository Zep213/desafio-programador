import csv
import io
import json

from openpyxl import load_workbook

from app.services.exporters import EXPORTERS

VALUE_CARTAO = {
    "pages": [
        {
            "page": 1,
            "days": [
                {
                    "date_raw": "01",
                    "punches": [
                        {"kind": "IN", "time_raw": "08:00", "time_hhmm": "08:00"},
                    ],
                },
                {
                    "date_raw": "02",
                    "punches": [
                        {"kind": "IN", "time_raw": "08:00", "time_hhmm": "08:00"},
                        {"kind": "OUT", "time_raw": "12:00", "time_hhmm": "12:00"},
                    ],
                },
            ],
        }
    ]
}

VALUE_HOLERITE = {
    "pages": [
        {
            "page": 1,
            "year": "2020",
            "month": "01",
            "fields": [
                {"code": "0010", "label": "Salário Base", "reference": "", "value": "1.000,00"},
            ],
            "bases": [{"label": "Total Vencimentos", "value": "1.000,00"}],
        },
        {"page": 1, "year": "2020", "month": "02", "fields": [], "bases": []},
    ]
}


def test_cartao_xlsx_cabecalho_e_destaque_amarelo_em_batida_impar():
    conteudo, mimetype = EXPORTERS["cartao-ponto"].export(VALUE_CARTAO, "xlsx")
    assert mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(io.BytesIO(conteudo))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Data", "Entrada 1", "Saída 1"]
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.color.rgb.endswith("FFFFFF")
    assert ws["A1"].fill.fgColor.rgb.endswith("173772")

    linha_impar = ws[2]
    assert linha_impar[0].value == "01"
    assert linha_impar[0].fill.fgColor.rgb.endswith("FFF3CD")


def test_cartao_xlsx_vermelho_ganha_quando_impar_e_nao_sequencial():
    value = {
        "pages": [
            {
                "page": 1,
                "days": [
                    {"date_raw": "05", "punches": []},
                    {
                        "date_raw": "03",
                        "punches": [{"kind": "IN", "time_raw": "08:00", "time_hhmm": "08:00"}],
                    },
                ],
            }
        ]
    }
    conteudo, _ = EXPORTERS["cartao-ponto"].export(value, "xlsx")
    wb = load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    linha_2 = ws[3]
    assert linha_2[0].value == "03"
    assert linha_2[0].fill.fgColor.rgb.endswith("F8D7DA")
    assert linha_2[0].border.left.color.rgb.endswith("DC3545")


def test_holerite_xlsx_colunas_e_pagina_vazia_amarela():
    conteudo, _ = EXPORTERS["holerite"].export(VALUE_HOLERITE, "xlsx")
    wb = load_workbook(io.BytesIO(conteudo))
    ws = wb.active

    assert [c.value for c in ws[1]] == ["Pág.", "Mês", "Ano", "Salário Base"]
    linha_vazia = ws[3]
    assert linha_vazia[0].fill.fgColor.rgb.endswith("FFF3CD")


def test_cartao_csv_usa_ponto_e_virgula_e_mesma_tabela_do_xlsx():
    conteudo, mimetype = EXPORTERS["cartao-ponto"].export(VALUE_CARTAO, "csv")
    assert mimetype == "text/csv"

    texto = conteudo.decode("utf-8-sig")
    linhas = list(csv.reader(io.StringIO(texto), delimiter=";"))
    assert linhas[0] == ["Data", "Entrada 1", "Saída 1"]
    assert linhas[1] == ["01", "08:00", ""]
    assert linhas[2] == ["02", "08:00", "12:00"]


def test_holerite_json_mesma_estrutura_tabular():
    conteudo, mimetype = EXPORTERS["holerite"].export(VALUE_HOLERITE, "json")
    assert mimetype == "application/json"

    corpo = json.loads(conteudo)
    assert corpo["colunas"] == ["Pág.", "Mês", "Ano", "Salário Base"]
    assert corpo["linhas"] == [
        {"Pág.": "1", "Mês": "01", "Ano": "2020", "Salário Base": "1.000,00"},
        {"Pág.": "1", "Mês": "02", "Ano": "2020", "Salário Base": ""},
    ]

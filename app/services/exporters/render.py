import csv
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app.services.tabela import Tabela, cor_da_linha

MIMETYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIMETYPE_CSV = "text/csv"
MIMETYPE_JSON = "application/json"

CSV_SEPARADOR = ";"

COR_CABECALHO = "173772"
COR_AMARELO = "FFF3CD"
COR_VERMELHO = "F8D7DA"
COR_BORDA_VERMELHA = "DC3545"

FONTE_CABECALHO = Font(bold=True, color="FFFFFF")
FUNDO_CABECALHO = PatternFill("solid", fgColor=COR_CABECALHO)
FUNDO_AMARELO = PatternFill("solid", fgColor=COR_AMARELO)
FUNDO_VERMELHO = PatternFill("solid", fgColor=COR_VERMELHO)
BORDA_ESQUERDA_VERMELHA = Border(left=Side(style="medium", color=COR_BORDA_VERMELHA))


def render_xlsx(tabela: Tabela) -> bytes:
    wb = Workbook()
    ws = wb.active

    ws.append(tabela.colunas)
    for celula in ws[1]:
        celula.font = FONTE_CABECALHO
        celula.fill = FUNDO_CABECALHO

    for linha in tabela.linhas:
        ws.append(linha.valores)
        cor = cor_da_linha(linha.avisos)
        if cor is None:
            continue
        linha_planilha = ws[ws.max_row]
        fundo = FUNDO_VERMELHO if cor == "vermelho" else FUNDO_AMARELO
        for celula in linha_planilha:
            celula.fill = fundo
        if cor == "vermelho":
            linha_planilha[0].border = BORDA_ESQUERDA_VERMELHA

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_csv(tabela: Tabela) -> bytes:
    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=CSV_SEPARADOR)
    escritor.writerow(tabela.colunas)
    for linha in tabela.linhas:
        escritor.writerow(linha.valores)
    return buffer.getvalue().encode("utf-8-sig")


def render_json(tabela: Tabela) -> bytes:
    linhas = [dict(zip(tabela.colunas, linha.valores)) for linha in tabela.linhas]
    conteudo = {"colunas": tabela.colunas, "linhas": linhas}
    return json.dumps(conteudo, ensure_ascii=False, indent=2).encode("utf-8")


RENDERIZADORES = {
    "xlsx": (render_xlsx, MIMETYPE_XLSX),
    "csv": (render_csv, MIMETYPE_CSV),
    "json": (render_json, MIMETYPE_JSON),
}
